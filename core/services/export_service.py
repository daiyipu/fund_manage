"""
评分报告导出服务
"""
import io
from typing import List, Dict, Optional
from datetime import datetime
from decimal import Decimal
import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from config.scoring_rules import SCORING_DIMENSIONS
from core.repositories.investment_repository import InvestmentRepository
from core.repositories.scoring_repository import ScoringRepository
from app.utils.database import get_db_connection

logger = logging.getLogger(__name__)


class ExportService:
    """评分报告导出服务"""

    def __init__(self):
        self.investment_repo = InvestmentRepository()
        self.scoring_repo = ScoringRepository()

    def export_scoring_report_excel(self, fund_id: int) -> bytes:
        """
        导出评分报告为Excel文件

        Args:
            fund_id: 投资ID

        Returns:
            Excel文件的字节流
        """
        try:
            # 获取投资信息
            investment = self.investment_repo.get_by_id(fund_id)
            if not investment:
                raise ValueError(f"投资 {fund_id} 不存在")

            # 获取评分详情
            scoring_detail = self._get_fund_scoring_detail(fund_id)

            # 创建工作簿
            wb = Workbook()
            wb.remove(wb.active)  # 删除默认sheet

            # 创建总览sheet
            self._create_overview_sheet(wb, investment, scoring_detail)

            # 创建各维度详情sheet
            for dim_code, dim_data in scoring_detail.get('dimensions', {}).items():
                self._create_dimension_sheet(wb, dim_code, dim_data, scoring_detail)

            # 保存到字节流
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()

        except Exception as e:
            logger.error(f"Error exporting scoring report: {str(e)}")
            raise

    def _get_fund_scoring_detail(self, fund_id: int) -> dict:
        """获取投资评分详情"""
        try:
            with get_db_connection() as conn:
                with conn.cursor() as cursor:
                    # 获取总分
                    sql = """
                        SELECT its.*, f.fund_name, f.fund_code
                        FROM fund_total_scores its
                        JOIN funds f ON its.fund_id = f.id
                        WHERE its.fund_id = %s
                    """
                    cursor.execute(sql, (fund_id,))
                    total_score = cursor.fetchone()

                    if not total_score:
                        return {'success': False, 'message': '未找到评分数据'}

                    # 获取各维度得分
                    sql = """
                        SELECT iss.*, sd.dimension_code, sd.dimension_name
                        FROM fund_scoring_summary iss
                        JOIN scoring_dimensions sd ON iss.dimension_id = sd.id
                        WHERE iss.fund_id = %s
                        ORDER BY sd.display_order
                    """
                    cursor.execute(sql, (fund_id,))
                    dimension_scores = cursor.fetchall()

                    # 获取指标得分
                    sql = """
                        SELECT ins.*,
                               si.indicator_code, si.indicator_name,
                               sd.dimension_code, sd.dimension_name
                        FROM fund_scores ins
                        JOIN scoring_indicators si ON ins.indicator_id = si.id
                        JOIN scoring_dimensions sd ON si.dimension_id = sd.id
                        WHERE ins.fund_id = %s
                        ORDER BY sd.display_order, si.display_order
                    """
                    cursor.execute(sql, (fund_id,))
                    indicator_scores = cursor.fetchall()

                    # 获取评分人信息
                    scorer_ids = list(set([s['scorer_id'] for s in indicator_scores]))
                    scorers = {}
                    if scorer_ids:
                        placeholders = ','.join(['%s'] * len(scorer_ids))
                        sql = f"SELECT id, real_name FROM users WHERE id IN ({placeholders})"
                        cursor.execute(sql, tuple(scorer_ids))
                        scorers = {u['id']: u['real_name'] for u in cursor.fetchall()}

                    # 整理数据
                    dimensions = {}
                    for dim in dimension_scores:
                        dim_code = dim['dimension_code']
                        dimensions[dim_code] = {
                            'name': dim['dimension_name'],
                            'total_score': float(dim['total_score']),
                            'weighted_total': float(dim['weighted_total']),
                            'indicators': []
                        }

                    for ind in indicator_scores:
                        dim_code = ind['dimension_code']
                        scorer_name = scorers.get(ind['scorer_id'], '未知')
                        dimensions[dim_code]['indicators'].append({
                            'code': ind['indicator_code'],
                            'name': ind['indicator_name'],
                            'score': float(ind['score']),
                            'weighted_score': float(ind['weighted_score']),
                            'scorer': scorer_name,
                            'comment': ind.get('scorer_comment', ''),
                            'scored_at': ind['scored_at'].strftime('%Y-%m-%d') if ind['scored_at'] else ''
                        })

                    return {
                        'success': True,
                        'data': {
                            'fund_name': total_score['fund_name'],
                            'fund_code': total_score['fund_code'],
                            'total_score': float(total_score['total_score']),
                            'policy_score': float(total_score['policy_score']),
                            'layout_score': float(total_score['layout_score']),
                            'execution_score': float(total_score['execution_score']),
                            'grade': total_score['grade'] or '-',
                            'rank': total_score['rank_in_period'],
                            'dimensions': dimensions
                        }
                    }
        except Exception as e:
            logger.error(f"Error getting fund scoring detail: {str(e)}")
            return {'success': False, 'message': str(e)}

    def _create_overview_sheet(self, wb: Workbook, fund: dict, scoring_detail: dict):
        """创建评分总览sheet"""
        ws = wb.create_sheet("评分总览", 0)

        # 标题
        ws['A1'] = '基金投向评分报告'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')

        # 基本信息
        row = 3
        ws[f'A{row}'] = '基金编码'
        ws[f'B{row}'] = scoring_detail.get('fund_code', '')
        row += 1
        ws[f'A{row}'] = '基金名称'
        ws[f'B{row}'] = scoring_detail.get('fund_name', '')
        row += 2

        # 评分结果
        ws[f'A{row}'] = '总分'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = scoring_detail.get('total_score', 0)
        ws[f'B{row}'].font = Font(bold=True, color="0066CC")
        row += 1
        ws[f'A{row}'] = '等级'
        ws[f'B{row}'] = scoring_detail.get('grade', '-')
        row += 1
        ws[f'A{row}'] = '排名'
        ws[f'B{row}'] = f"第 {scoring_detail.get('rank', '-')} 名" if scoring_detail.get('rank') else '-'

        # 空行
        row += 2

        # 评分结果
        ws[f'A{row}'] = '总分'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = scoring_detail.get('total_score', 0)
        ws[f'B{row}'].font = Font(bold=True, color="0066CC")
        row += 1
        ws[f'A{row}'] = '等级'
        ws[f'B{row}'] = scoring_detail.get('grade', '-')
        row += 1
        ws[f'A{row}'] = '排名'
        ws[f'B{row}'] = f"第 {scoring_detail.get('rank', '-')} 名" if scoring_detail.get('rank') else '-'

        # 列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30

        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for r in range(3, row + 1):
            for c in ['A', 'B']:
                ws[f'{c}{r}'].border = thin_border
                ws[f'{c}{r}'].alignment = Alignment(horizontal='left', vertical='center')

    def _create_dimension_sheet(self, wb: Workbook, dim_code: str, dim_data: dict, scoring_detail: dict):
        """创建维度详情sheet"""
        # 从 SCORING_DIMENSIONS 获取维度名称
        dim_config = SCORING_DIMENSIONS.get(dim_code, {})
        dim_name = dim_config.get('name', dim_code)

        ws = wb.create_sheet(dim_name)

        # 构建层级嵌套数据
        rows = self._build_dimension_rows(dim_code, dim_data, dim_config)

        # 表头
        headers = ['维度', '指标', '子指标', '得分', '满分', '权重(%)', '加权得分', '评分人', '评分时间']
        ws.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 数据行
        for row_data in rows:
            ws.append(row_data)

        # 列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15

        # 设置边框和对齐
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def _build_dimension_rows(self, dim_code: str, dim_data: dict, dim_config: dict) -> List[list]:
        """构建维度数据的层级嵌套行"""
        rows = []
        dim_name = dim_config.get('name', dim_code)

        # 遍历指标
        for indicator in dim_config.get('indicators', []):
            ind_code = indicator['code']
            ind_name = indicator['name']
            ind_max_score = indicator.get('max_score', 0)

            # 判断是叶子指标还是父指标
            if indicator.get('type') == 'leaf':
                # 叶子指标，直接评分
                # 从评分数据中找到对应的分数
                score_data = next(
                    (s for s in dim_data.get('indicators', []) if s['code'] == ind_code),
                    None
                )
                if score_data:
                    rows.append([
                        dim_name,
                        ind_name,
                        '-',
                        score_data['score'],
                        ind_max_score,
                        f"{indicator.get('weight', 0):.2f}",
                        score_data['weighted_score'],
                        score_data['scorer'],
                        score_data['scored_at']
                    ])

            elif indicator.get('type') == 'parent':
                # 父指标，有子指标
                sub_indicators = indicator.get('sub_indicators', [])
                subtotal_score = 0
                subtotal_max = ind_max_score

                for sub_ind in sub_indicators:
                    sub_code = sub_ind['code']
                    sub_name = sub_ind['name']
                    sub_max = sub_ind.get('max_score', 0)

                    score_data = next(
                        (s for s in dim_data.get('indicators', []) if s['code'] == sub_code),
                        None
                    )
                    if score_data:
                        subtotal_score += score_data['score']
                        rows.append([
                            dim_name,
                            ind_name,
                            sub_name,
                            score_data['score'],
                            sub_max,
                            f"{sub_ind.get('weight', 0):.2f}",
                            score_data['weighted_score'],
                            score_data['scorer'],
                            score_data['scored_at']
                        ])

                # 添加小计行
                if subtotal_score > 0:
                    rows.append([
                        dim_name,
                        ind_name,
                        '**小计**',
                        subtotal_score,
                        subtotal_max,
                        f"{indicator.get('weight', 0):.2f}",
                        f"{subtotal_score * (indicator.get('weight', 0) / 100):.2f}",
                        '-',
                        '-'
                    ])

        # 添加维度合计
        dim_total = dim_data.get('total_score', 0)
        dim_max = dim_config.get('max_score', 0)
        dim_weight = dim_config.get('weight', 0)
        rows.append([
            f'**{dim_name}合计**',
            '-',
            '-',
            dim_total,
            dim_max,
            f"{dim_weight:.2f}",
            dim_data.get('weighted_total', 0),
            '-',
            '-'
        ])

        return rows


# 创建全局实例
export_service = ExportService()
